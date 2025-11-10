"""
报损数据统计系统 - 多店铺Excel版
核心功能：支持多店铺独立统计，生成带店铺名称的Excel报告
"""
import requests
import re
from datetime import datetime, timedelta
import time
from bs4 import BeautifulSoup
import logging
from typing import Dict, List, Optional, Tuple, Union
from dataclasses import dataclass, field
from abc import ABCMeta, abstractmethod
# Excel 相关依赖
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ------------------------------
# 1. 配置层（新增多店铺配置）
# ------------------------------
@dataclass
class Config:
    base_url: str = "http://www.mspic.cn/admin/replenish/order/bs"
    login_url: str = "http://www.mspic.cn/admin/login?fromuri=/admin/home"
    # 新增：多店铺配置（店铺ID: 店铺名称）
    shops: Dict[str, str] = field(default_factory=lambda: {
        "1742702902722205": "芙蓉街",
        "1742702902722200": "宽厚里",
        "1713879145388131": "台东",
        "1734490564086185": "市南",
        "1584690317016138": "胶州"
    })
    username: str = "yuanhao"
    password: str = "086618"
    target_year: int = 2025
    target_month: int = 9
    request_timeout: int = 15
    request_interval: float = 1.5
    retry_count: int = 3
    retry_delay: float = 2.0
    parse_keywords: List[str] = field(default_factory=lambda: ["报损金额", "金额", "报损", "报损详情展示"])
    currency_symbols: List[str] = field(default_factory=lambda: ["¥", "￥", ",", "【", "】", "(", ")"])
    log_file: str = "baosun_log.txt"
    report_file_prefix: str = "baosun_report_"
    encoding: str = "utf-8"
    debug_mode: bool = True
    report_format: str = "excel"

# ------------------------------
# 2. 工具层（无修改）
# ------------------------------
class Toolkit:
    @staticmethod
    def generate_dates(year: int, month: int) -> List[str]:
        dates = []
        try:
            current_date = datetime(year, month, 1)
            while current_date.month == month:
                dates.append(current_date.strftime("%Y-%m-%d"))
                current_date += timedelta(days=1)
        except ValueError as e:
            logging.error(f"生成日期失败：{e}")
        return dates
    
    @staticmethod
    def clean_amount(text: str, currency_symbols: List[str]) -> float:
        if not text:
            return 0.0
        clean_text = text.strip()
        for symbol in currency_symbols:
            clean_text = clean_text.replace(symbol, "")
        if clean_text and clean_text.replace(".", "", 1).isdigit():
            return round(float(clean_text), 2)
        return 0.0
    
    @staticmethod
    def init_logger(config: Config):
        logging.basicConfig(
            level=logging.DEBUG if config.debug_mode else logging.INFO,
            format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
            handlers=[
                logging.StreamHandler(),
                logging.FileHandler(config.log_file, encoding=config.encoding, mode="a")
            ]
        )
        return logging.getLogger(__name__)
    
    @staticmethod
    def extract_csrf_token(html_content: str) -> Optional[Dict[str, str]]:
        soup = BeautifulSoup(html_content, "html.parser")
        csrf_data = {}
        csrf_fields = ["_token", "csrf_token", "csrfmiddlewaretoken"]
        for field in csrf_fields:
            input_tag = soup.find("input", attrs={"type": "hidden", "name": field})
            if input_tag and "value" in input_tag.attrs:
                csrf_data[field] = input_tag["value"]
                logging.debug(f"提取到CSRF Token：{field}={input_tag['value'][:10]}...")
                return csrf_data
        hidden_inputs = soup.find_all("input", attrs={"type": "hidden"})
        for input_tag in hidden_inputs:
            if "name" in input_tag.attrs and "value" in input_tag.attrs:
                csrf_data[input_tag["name"]] = input_tag["value"]
        if csrf_data:
            logging.debug(f"提取到隐藏字段：{list(csrf_data.keys())}")
            return csrf_data
        logging.warning("未提取到CSRF Token，可能导致登录失败")
        return None
    
    @staticmethod
    def extract_amount_by_regex(text: str) -> float:
        # 增强版正则：支持换行、多空格、中英文括号
        pattern = r"报损详情展示[\s\S]*?(?:【|\()(\d+\.?\d*)(?:】|\))"
        matches = re.findall(pattern, text, re.IGNORECASE)
        
        if matches:
            total = 0.0
            for amount_text in matches:
                amount_text = amount_text.strip()
                if amount_text and amount_text.replace(".", "", 1).isdigit():
                    amount = round(float(amount_text), 2)
                    total += amount
                    logging.debug(f"正则匹配到金额：{amount_text}（累计：{total}）")
            return total
        
        if "报损详情展示" in text:
            idx = text.find("报损详情展示")
            debug_text = text[max(0, idx-20):min(len(text), idx+100)]
            logging.debug(f"未匹配到金额！文本片段：{debug_text.replace(chr(10), '\\n').replace(chr(13), '\\r')}")
        else:
            logging.debug("未找到“报损详情展示”关键词")
        
        return 0.0

# ------------------------------
# 3. 核心服务层（无修改）
# ------------------------------
class BaseLoginService(metaclass=ABCMeta):
    @abstractmethod
    def login(self) -> bool:
        pass

class BaseDataFetcher(metaclass=ABCMeta):
    @abstractmethod
    def fetch(self, date_str: str) -> Optional[str]:
        pass

class BaseDataParser(metaclass=ABCMeta):
    @abstractmethod
    def parse(self, html_content: str, date_str: str) -> float:
        pass

class BaseReportService(metaclass=ABCMeta):
    @abstractmethod
    def generate(self, daily_data: Dict[str, float], total_amount: float, shop_name: str) -> None:
        # 新增 shop_name 参数：用于生成带店铺名称的报告
        pass

# ------------------------------
# 4. 具体实现类（重点：多店铺适配）
# ------------------------------
class MspicLoginService(BaseLoginService):
    def __init__(self, config: Config, session: requests.Session):
        self.config = config
        self.session = session
        self.logged_in = False
        self.session.headers.update({
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/129.0.0.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
            "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
            "Referer": self.config.login_url,
            "Content-Type": "application/x-www-form-urlencoded",
        })
    
    def login(self) -> bool:
        try:
            logging.info(f"访问登录页：{self.config.login_url}")
            login_page_response = self.session.get(
                self.config.login_url,
                timeout=self.config.request_timeout
            )
            login_page_response.raise_for_status()
            login_page_url = login_page_response.url
            
            csrf_data = Toolkit.extract_csrf_token(login_page_response.text)
            login_data = {
                "username": self.config.username,
                "password": self.config.password,
                "remember": "on"
            }
            if csrf_data:
                login_data.update(csrf_data)
            
            logging.info("提交登录请求...")
            response = self.session.post(
                self.config.login_url,
                data=login_data,
                timeout=self.config.request_timeout,
                allow_redirects=True
            )
            
            if self._is_login_success(response, login_page_url):
                self.logged_in = True
                logging.info(f"登录成功！当前URL：{response.url}")
                print("✅ 登录成功")
                return True
            else:
                response_text = response.text[:500].replace("\n", " ")
                logging.error(f"登录失败！响应内容前500字符：{response_text}")
                logging.error(f"登录后URL：{response.url}")
                print(f"❌ 登录失败，详情查看日志文件")
                return False
                
        except requests.exceptions.ConnectionError:
            logging.error("登录失败：无法连接服务器，请检查网络")
            print("❌ 登录失败：无法连接服务器")
            return False
        except requests.exceptions.Timeout:
            logging.error("登录失败：请求超时")
            print("❌ 登录失败：请求超时")
            return False
        except requests.exceptions.HTTPError as e:
            logging.error(f"登录失败：HTTP错误 {e}")
            print(f"❌ 登录失败：HTTP错误 {e}")
            return False
        except Exception as e:
            logging.error(f"登录失败：未知错误 -> {str(e)}", exc_info=True)
            print(f"❌ 登录失败：未知错误 {e}")
            return False
    
    def _is_login_success(self, response: requests.Response, login_page_url: str) -> bool:
        success_conditions = [
            response.url != login_page_url and "/admin/" in response.url,
            "欢迎" in response.text or "控制台" in response.text or "数据统计" in response.text,
            "用户名或密码错误" not in response.text and "登录" not in response.text[:200],
            response.status_code in [200, 302]
        ]
        return sum(success_conditions) >= 2

class MspicDataFetcher(BaseDataFetcher):
    def __init__(self, config: Config, session: requests.Session, login_service: BaseLoginService, shop_id: str):
        self.config = config
        self.session = session
        self.login_service = login_service
        self.shop_id = shop_id  # 每个Fetcher绑定一个店铺ID
    
    def fetch(self, date_str: str) -> Optional[str]:
        if not self.login_service.logged_in:
            logging.error(f"店铺{self.shop_id} - 获取{date_str}数据失败：未登录")
            return None
        
        params = {
            "shopViewId": self.shop_id,  # 使用当前店铺ID
            "day": date_str
        }
        
        for retry in range(self.config.retry_count):
            try:
                response = self.session.get(
                    self.config.base_url,
                    params=params,
                    timeout=self.config.request_timeout
                )
                
                if response.status_code == 200:
                    response.encoding = self.config.encoding
                    logging.info(f"店铺{self.shop_id} - 成功获取{date_str}数据（编码：{response.encoding}）")
                    return response.text
                elif response.status_code == 503:
                    logging.warning(f"店铺{self.shop_id} - 获取{date_str}数据失败：服务器暂时不可用，重试第{retry+1}次")
                    time.sleep(self.config.retry_delay)
                elif response.status_code == 302:
                    logging.error(f"店铺{self.shop_id} - 获取{date_str}数据失败：登录状态失效")
                    return None
                else:
                    logging.error(f"店铺{self.shop_id} - 获取{date_str}数据失败：状态码{response.status_code}")
                    return None
            except Exception as e:
                logging.warning(f"店铺{self.shop_id} - 获取{date_str}数据失败：重试第{retry+1}次 -> {str(e)}")
                time.sleep(self.config.retry_delay)
        
        logging.error(f"店铺{self.shop_id} - 获取{date_str}数据失败：已达最大重试次数")
        return None

class MspicDataParser(BaseDataParser):
    def __init__(self, config: Config, shop_id: str):
        self.config = config
        self.shop_id = shop_id  # 每个Parser绑定一个店铺ID
    
    def parse(self, html_content: str, date_str: str) -> float:
        total_amount = 0.0
        
        try:
            if self.config.debug_mode:
                target_keywords = ["报损详情展示", "【", "】", "(", ")"]
                for keyword in target_keywords:
                    count = html_content.count(keyword)
                    logging.debug(f"店铺{self.shop_id} - {date_str} - “{keyword}”出现次数：{count}")
            
            # 策略0：正则匹配
            total_amount = Toolkit.extract_amount_by_regex(html_content)
            if total_amount > 0:
                logging.info(f"店铺{self.shop_id} - {date_str} 正则匹配金额：¥{total_amount}")
                return total_amount
            
            # 策略1：关键词匹配
            total_amount = self._parse_by_keyword(html_content)
            if total_amount > 0:
                logging.info(f"店铺{self.shop_id} - {date_str} 关键词匹配金额：¥{total_amount}")
                return total_amount
            
            # 策略2：表格解析
            total_amount = self._parse_by_table(html_content)
            if total_amount == 0:
                logging.warning(f"店铺{self.shop_id} - {date_str} 未解析到报损数据")
            else:
                logging.info(f"店铺{self.shop_id} - {date_str} 表格解析金额：¥{total_amount}")
            
            return total_amount
        except Exception as e:
            logging.error(f"店铺{self.shop_id} - {date_str} 解析失败：{str(e)}", exc_info=True)
            return 0.0
    
    def _parse_by_keyword(self, html_content: str) -> float:
        soup = BeautifulSoup(html_content, "html.parser")
        for keyword in self.config.parse_keywords:
            rows = soup.find_all("tr", string=lambda text: text and keyword in text)
            for row in rows:
                cells = row.find_all("td")
                if len(cells) >= 2:
                    return Toolkit.clean_amount(cells[1].get_text(), self.config.currency_symbols)
        return 0.0
    
    def _parse_by_table(self, html_content: str) -> float:
        soup = BeautifulSoup(html_content, "html.parser")
        tables = soup.find_all("table")
        total_amount = 0.0
        
        for table in tables:
            amount_col_index = self._find_amount_column(table)
            if amount_col_index == -1:
                continue
            
            data_rows = table.find_all("tr")[1:]
            for row in data_rows:
                cells = row.find_all("td")
                if len(cells) > amount_col_index:
                    amount_text = cells[amount_col_index].get_text().strip()
                    total_amount += Toolkit.clean_amount(amount_text, self.config.currency_symbols)
        
        return total_amount
    
    def _find_amount_column(self, table: BeautifulSoup) -> int:
        headers = table.find_all("th")
        for idx, header in enumerate(headers):
            header_text = header.get_text().strip()
            if any(keyword in header_text for keyword in self.config.parse_keywords):
                return idx
        return -1

class ExcelReportService(BaseReportService):
    """Excel 报告生成服务（支持多店铺，带店铺名称）"""
    def __init__(self, config: Config):
        self.config = config
    
    def _generate_report_filename(self, shop_name: str) -> str:
        """生成带店铺名称的Excel文件名"""
        date_str = f"{self.config.target_year}{self.config.target_month:02d}"
        return f"{self.config.report_file_prefix}{shop_name}_{date_str}.xlsx"
    
    def _set_excel_styles(self, ws):
        """设置Excel样式（和之前一致）"""
        # 表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 边框样式
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # 应用表头样式
        for col in range(1, 4):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 应用数据行样式
        for row in range(2, ws.max_row + 1):
            for col in range(1, 4):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if col in [2, 3]:
                    cell.number_format = "0.00"
        
        # 合计行样式
        total_row = ws.max_row
        total_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        for col in range(1, 4):
            cell = ws.cell(row=total_row, column=col)
            cell.font = Font(bold=True)
            cell.fill = total_fill
    
    def _auto_adjust_column_width(self, ws):
        """自动调整列宽"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def generate(self, daily_data: Dict[str, float], total_amount: float, shop_name: str) -> None:
        """生成带店铺名称的Excel报告"""
        self._print_to_console(daily_data, total_amount, shop_name)
        
        try:
            # 创建新的工作簿（每个店铺独立工作簿）
            wb = Workbook()
            ws = wb.active
            # 工作表名称：店铺名称+年月
            ws.title = f"{shop_name}_{self.config.target_year}年{self.config.target_month}月报损统计"
            
            # 写入表头
            headers = ["统计时间", "每日报损金额（元）", "月度合计金额（元）"]
            ws.append(headers)
            
            # 写入每日数据
            for date, amount in sorted(daily_data.items()):
                ws.append([date, amount, ""])
            
            # 写入合计行
            ws.append(["合计", "", total_amount])
            
            # 设置样式和列宽
            self._set_excel_styles(ws)
            self._auto_adjust_column_width(ws)
            
            # 保存文件（带店铺名称）
            report_file = self._generate_report_filename(shop_name)
            wb.save(report_file)
            
            logging.info(f"店铺{shop_name} - Excel报告已保存到：{report_file}")
            print(f"\n💾 {shop_name} - Excel报告保存路径：{report_file}")
        
        except Exception as e:
            logging.error(f"店铺{shop_name} - 保存Excel报告失败：{str(e)}", exc_info=True)
            print(f"❌ {shop_name} - 保存Excel报告失败：{e}")
    
    def _print_to_console(self, daily_data: Dict[str, float], total_amount: float, shop_name: str) -> None:
        """控制台输出带店铺名称"""
        print(f"\n{'='*80}")
        print(f"📊 {shop_name} - {self.config.target_year}年{self.config.target_month}月报损数据统计")
        print(f"{'='*80}")
        print(f"{'统计时间':<15} {'金额':<15} {'合计金额':<15}")
        print("-"*80)
        for date, amount in sorted(daily_data.items()):
            print(f"{date:<15} ¥{amount:>13.2f} {'':<15}")
        print("-"*80)
        print(f"{'合计':<15} {'':<15} ¥{total_amount:>13.2f}")
        print(f"{'='*80}")

# ------------------------------
# 5. 应用入口（重点：多店铺循环处理）
# ------------------------------
class BaoSunApplication:
    def __init__(self, config: Optional[Config] = None):
        self.config = config or Config()
        self.logger = Toolkit.init_logger(self.config)
        self.total_summary: Dict[str, float] = {}  # 存储所有店铺的总金额汇总
    
    def run_single_shop(self, shop_id: str, shop_name: str) -> Tuple[float, Dict[str, float]]:
        """处理单个店铺的统计"""
        print(f"\n{'='*80}")
        print(f"🚀 开始处理店铺：{shop_name}（店铺ID：{shop_id}）")
        print(f"{'='*80}")
        
        # 每个店铺独立创建Session（避免登录状态冲突）
        session = requests.Session()
        login_service = MspicLoginService(self.config, session)
        data_fetcher = MspicDataFetcher(self.config, session, login_service, shop_id)
        data_parser = MspicDataParser(self.config, shop_id)
        report_service = ExcelReportService(self.config)
        
        total_amount = 0.0
        daily_data: Dict[str, float] = {}
        
        try:
            # 登录（每个店铺独立登录，确保数据隔离）
            if not login_service.login():
                logging.error(f"店铺{shop_name} - 登录失败，跳过该店铺")
                print(f"❌ {shop_name} - 登录失败，跳过")
                return 0.0, {}
            
            # 生成目标月份的日期
            dates = Toolkit.generate_dates(self.config.target_year, self.config.target_month)
            if not dates:
                logging.error(f"店铺{shop_name} - 生成日期失败，跳过该店铺")
                print(f"❌ {shop_name} - 生成日期失败，跳过")
                return 0.0, {}
            
            # 循环处理每日数据
            for idx, date_str in enumerate(dates, 1):
                print(f"\n[{idx}/{len(dates)}] 处理 {date_str}...")
                html_content = data_fetcher.fetch(date_str)
                if not html_content:
                    daily_data[date_str] = 0.0
                    print(f"{date_str}：❌ 获取数据失败")
                    continue
                
                amount = data_parser.parse(html_content, date_str)
                daily_data[date_str] = amount
                total_amount += amount
                print(f"{date_str}：✅ 金额¥{amount:.2f}")
                
                if idx < len(dates):
                    time.sleep(self.config.request_interval)
            
            # 生成该店铺的Excel报告
            report_service.generate(daily_data, total_amount, shop_name)
            self.total_summary[shop_name] = total_amount
            logging.info(f"店铺{shop_name} - 统计完成！总金额：¥{total_amount:.2f}")
            return total_amount, daily_data
        
        except KeyboardInterrupt:
            logging.info(f"店铺{shop_name} - 程序被用户中断")
            print(f"\n⏹️ {shop_name} - 程序中断")
            return total_amount, daily_data
        except Exception as e:
            logging.error(f"店铺{shop_name} - 执行失败：{str(e)}", exc_info=True)
            print(f"\n❌ {shop_name} - 执行失败：{e}")
            return total_amount, daily_data
        finally:
            session.close()
            logging.info(f"店铺{shop_name} - HTTP会话已关闭")
    
    def run(self) -> Dict[str, float]:
        """主运行入口：循环处理所有店铺"""
        logging.info("="*50)
        logging.info(f"开始执行多店铺报损统计（目标年月：{self.config.target_year}年{self.config.target_month}月）")
        logging.info(f"待处理店铺：{list(self.config.shops.values())}")
        logging.info("="*50)
        
        # 循环处理每个店铺
        for shop_id, shop_name in self.config.shops.items():
            self.run_single_shop(shop_id, shop_name)
        
        # 输出所有店铺的汇总信息
        self.print_all_shops_summary()
        return self.total_summary
    
    def print_all_shops_summary(self):
        """输出所有店铺的总金额汇总"""
        print(f"\n{'='*80}")
        print(f"📋 所有店铺{self.config.target_year}年{self.config.target_month}月汇总")
        print(f"{'='*80}")
        grand_total = 0.0
        for shop_name, total_amount in self.total_summary.items():
            print(f"{shop_name:<15} ¥{total_amount:>15.2f}")
            grand_total += total_amount
        print("-"*80)
        print(f"{'所有店铺合计':<15} ¥{grand_total:>15.2f}")
        print(f"{'='*80}")
        logging.info(f"所有店铺汇总完成！总金额：¥{grand_total:.2f}")

# ------------------------------
# 6. 执行入口（无修改）
# ------------------------------
def main():
    app = BaoSunApplication()
    app.run()

if __name__ == "__main__":
    main()